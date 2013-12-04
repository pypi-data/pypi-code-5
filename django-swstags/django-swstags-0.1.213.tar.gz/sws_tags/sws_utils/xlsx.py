#!/usr/bin/env python
# encoding: utf-8

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from sws_tags.sws_utils import json_encode

import csv
from django.http import HttpResponse
from django.template import loader, Context
from decimal import Decimal
from datetime import date, datetime, timedelta
import time
import ujson
# import datetime



# FOR EXCEL GENERATING
# import xlwt
# ezxf = xlwt.easyxf

from sws_tags.sws_utils.common_utils import *
from sws_tags.sws_utils.messages import *
from sws_tags.sws_utils.cube import *

import traceback



# FOR EXCEL 
try:
	import cStringIO as StringIO
except ImportError:
	import StringIO

from django.http import HttpResponse

from xlsxwriter.workbook import Workbook




# FOR EXCEL GENERATING
import xlwt
ezxf = xlwt.easyxf

from sws_tags.sws_utils.common_utils import *
from sws_tags.sws_utils.messages import *
from sws_tags.sws_utils.cube import *

import traceback


import codecs

# import xlwt workbook = xlwt.Workbook(encoding = 'ascii') 
# worksheet = workbook.add_sheet('My Worksheet') 
# worksheet.write(0, 0, label = 'Row 0, Column 0 Value') 
# workbook.save('Excel_Workbook.xls') 

def ExportFile(param_export_file):
	def read_and_flush():
		output.seek(0)
		data = output.read()
		output.seek(0)
		output.truncate()
		return data

	def writeInStream():
			row = 0
			col = 0
			for h in headers:
				sheet.write(row, col, h,bold)
				col += 1
			row = 1
			col = 0
			for trr in queryset:
				for c in col_name:
					if type(trr[c]) == datetime:
						try:
							trr[c]=param_export_file['request_data']['django_timezone'].normalize(trr[c]).strftime('%Y-%m-%d %H:%M:%S')
						except:
							# print traceback.format_exc()
							trr[c]=str(trr[c])
					sheet.write(row, col, trr[c])
					col += 1
				row += 1
				col = 0
				data = read_and_flush()
				yield data

			book.close()

			data = read_and_flush()
			yield data	


	def writeBillInStream():
		bold = book.add_format({'bold': 1})
		border = book.add_format({'bold': True,'font_color': 'green','border_color':'blue','border': 6})
		font_color_red = book.add_format({'bold': True,'color':'red'})
		yellow_bg = book.add_format({'bg_color':'#FBE067'})
		grey_bg = book.add_format({'bg_color':'#D7D6D2'})
		grey_dark_bg_bold = book.add_format({'bg_color':'#656462','bold':1})

		flesh_color_bg = book.add_format({'bg_color':'#F6D4B2'})
		flesh_color_bg_bold = book.add_format({'bg_color':'#F6D4B2','bold':1})



		merge_format_border = book.add_format({
		    'bold': 1,
		    'border': 1,
		    'border_color':'#8E0172',
		    'align': 'center',
		    'valign': 'vcenter',
		    'fg_color': '#F6D4B2'})

		merge_format = book.add_format({
		    'bold': 1,
		    'align': 'center',
		    'valign': 'vcenter',
		    'fg_color': '#F6D4B2'})




		money = book.add_format({'num_format': '#.###'+param_export_file['data_bill']['type_money'],'bold':1})
		date_format = book.add_format({'num_format': 'mmmm d yyyy'})
		# italic = book.add_format({'italic': True})

		pattern = xlwt.Pattern() # Create the Pattern 
		pattern.pattern = xlwt.Pattern.SOLID_PATTERN
		pattern_fore_colour = 5

		style = xlwt.XFStyle() # Create the Pattern 
		style.pattern = pattern 
		


		# param_export_file['raw_cdrts'] 
		# param_export_file['cdrt_subtotal']
		# {'sum_total_income': Decimal('1.4883'),
		#  'sum_total_minutes': 9.8166666666666593, 
		#  'providerrange__providerdestination__destination': u'NICARAGUA MOBILE',
		#   'income': Decimal('0.1516')},




		# queryset =OrderedDict([('Destination', 'Afghanistan'), ('Total Minutes', 523), ('EUR/min', 0.35), ('Total', 183.05)]), OrderedDict([('Destination', 'Mexico'), ('Total Minutes', 837), ('EUR/min', 0.4), ('Total', 334.8)]),OrderedDict([('Destination', 'Marruecos'), ('Total Minutes', 100), ('EUR/min', 0.35), ('Total', 35.0)])


		queryset = param_export_file['data_bill']['raw_cdrts'] 
		total = param_export_file['data_bill']['cdrt_subtotal']



		# headers = ['providerrange__providerdestination__destination','sum_total_minutes','income','sum_total_income']
		col_name = ['providerrange__providerdestination__destination','sum_total_minutes','income','sum_total_income']
		headers = ['Destination','Total Minutes','EUR/min','Total']

		try:
			sheet.insert_image('A1',param_export_file['data_bill']['url_img'],{'x_offset':15,'y_offset':15})
			sheet.write(0,2,param_export_file['data_bill']['issuing_company_name'],bold)
			sheet.write(1,2,param_export_file['data_bill']['issuing_company_address'])
			sheet.write(3,2,param_export_file['data_bill']['issuing_company_CIF'])
			sheet.write(0,4,'FACTURA',font_color_red)
			sheet.write(1,4,param_export_file['data_bill']['invoice_number'],bold)

			# url =  param_export_file['STATIC_URL'] +'/imgcompany/sultan_logo.png'
			# sheet.insert_image('A1',url,{'x_offset':15,'y_offset':15})
			# sheet.write(0,2,'SULTAN',bold)
			# sheet.write(1,2,'Alfonso Gomez 30 4 planta')
			# sheet.write(3,2,'CIF: 698787651231')
			# sheet.write(0,4,'FACTURA',font_color_red)
			# sheet.write(1,4,'Numero: 132123',bold)
			# sheet.set_column('A:A', 20)




		except:
			print traceback.format_exc()
		
		
		row = 5
		col = 1
		for h in headers:
			sheet.write(row, col, h,grey_dark_bg_bold)
			col += 1
		row = 6
		col = 1

		col_max=len(col_name)



		for trr in queryset:
			for c in col_name:
				if type(trr[c]) == datetime:
					try:
						trr[c]=param_export_file['request_data']['django_timezone'].normalize(trr[c]).strftime('%Y-%m-%d %H:%M:%S')
					except:
						# print traceback.format_exc()
						trr[c]=str(trr[c])
				if row%2 ==0:
					sheet.write(row, col, trr[c],yellow_bg)
				else:
					sheet.write(row,col,trr[c],grey_bg)

				col += 1
			row += 1
			col = 1
			data = read_and_flush()
			yield data

		sheet.write(row, col_max-3, 'Total Minutes',bold)	
		sheet.write(row, col_max-2, total['sum_total_minutes'],bold)	
		sheet.write(row, col_max-1, 'Total',bold)
		sheet.write(row, col_max, total['sum_total_income'],money)	

		row = row+1
		sheet.write(row, col_max-2,'IVA')
		sheet.write(row, col_max-1,'21 %')
		total_iva = float(total['sum_total_income']) * 1.21
		sheet.write(row, col_max,total_iva,money)


		# row = row +2
		# sheet.write(row, col_max-3,'Period From:',bold)
		# sheet.write(row,col_max-2,param_export_file['data_bill']['from_date'])		
		# sheet.write(row,col_max-1,'Period To:',bold)
		# sheet.write(row,col_max,param_export_file['data_bill']['to_date'])


		# row = row +2
		# sheet.write(row, col_max-3,'Issue Date:',bold)
		# sheet.write(row,col_max-2,param_export_file['data_bill']['issue_date'])		
		# sheet.write(row,col_max-1,'Due Date:',bold)
		# sheet.write(row,col_max,param_export_file['data_bill']['due_date'])


		# Merge 3 cells.


		row = row +3
		
		# print 'row-->',row
		# print 'col-->',col_max

		coordinate = 'B'+str(row)+':E'+str(row)
		period = 'Issue Date: '+ str(param_export_file['data_bill']['issue_date'])+ '      Due Date: '+str(param_export_file['data_bill']['due_date'])
		sheet.merge_range(coordinate, period, merge_format_border)

		row = row +2

		coordinate = 'B'+str(row)+':E'+str(row)
		period = 'Period From: '+ str(param_export_file['data_bill']['from_date'])+ '      Period To: '+str(param_export_file['data_bill']['to_date'])
		sheet.merge_range(coordinate, period, merge_format_border)


		row = row +2

		coordinate = 'B'+str(row)+':E'+str(row)
		period = 'Payment info for client: '+ param_export_file['data_bill']['payment'] 
		sheet.merge_range(coordinate, period, merge_format)

		row = row-1

		sheet.write(row+1, 1, 'Bank Name:',flesh_color_bg_bold)	
		sheet.write(row+2, 1, 'Account:',flesh_color_bg_bold)
		sheet.write(row+3, 1, 'Beneficiary:',flesh_color_bg_bold)
		sheet.write(row+4, 1, 'Bank addres:',flesh_color_bg_bold)

		sheet.write(row+1, 2, param_export_file['data_bill']['bank_name'],flesh_color_bg)	
		sheet.write(row+2, 2, param_export_file['data_bill']['account'],flesh_color_bg)
		sheet.write(row+3, 2, param_export_file['data_bill']['beneficiary'],flesh_color_bg)
		sheet.write(row+4, 2, param_export_file['data_bill']['bank_addres'],flesh_color_bg)


		sheet.write(row+1, 3, 'IBAN:',flesh_color_bg_bold)	
		sheet.write(row+2, 3, 'SWIFT:',flesh_color_bg_bold)
		sheet.write(row+3, 3, 'Company:',flesh_color_bg_bold)
		sheet.write(row+4, 3, 'Company Number:',flesh_color_bg_bold)
		sheet.write(row+5, 3, 'Email:',flesh_color_bg_bold)
		
		sheet.write(row+1, 4, param_export_file['data_bill']['IBAN'],flesh_color_bg)	
		sheet.write(row+2, 4, param_export_file['data_bill']['SWIFT'],flesh_color_bg)
		sheet.write(row+3, 4, param_export_file['data_bill']['company'],flesh_color_bg)
		sheet.write(row+4, 4, param_export_file['data_bill']['company_number'],flesh_color_bg)
		sheet.write(row+5, 4, param_export_file['data_bill']['email'],flesh_color_bg)


		sheet.set_column('B:B', 15)
		sheet.set_column('C:C', 15)
		sheet.set_column('D:D', 15)
		sheet.set_column('E:E', 15)		


		sheet.set_column('G:G', 15)

		# marge_format = book.add_format({'align':center})

		# sheet.merge_range('B3:D4','Merged Cell',merge_format_border)


		# sheet.write('A3', 'Cell A3', format)


		book.close()

		data = read_and_flush()
		yield data	


	# print 'AAAAAAAAAAAAAAAAAAAAAA'
	# print param_export_file


	if param_export_file['format'] == 'excel':
		param_export_file['filename'] = param_export_file['filename']+'_' + str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))+'.xlsx'

		queryset = param_export_file['queryset']
		headers = param_export_file['headers']
		col_name = param_export_file['col_name']
		filename = param_export_file['filename']
		logger = param_export_file['logger']

		# print queryset

		output = StringIO.StringIO()

		book = Workbook(output)
		sheet = book.add_worksheet('test') 

		bold = book.add_format({'bold': 1})

		response = HttpResponse(writeInStream(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
		response['Content-Disposition'] = "attachment; filename="+ filename
		return response

	elif param_export_file['format'] == 'csv':
		param_export_file['filename'] = param_export_file['filename']+'_'+str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))

		json_data = get_csv_query(param_export_file['request_data'], param_export_file['queryset'], param_export_file['filename'])
		response = HttpResponse(json_data, mimetype='text/csv') 
		param_export_file['filename'] += '.csv'
		response['Content-Disposition'] = 'attachment; filename=' + param_export_file['filename']
		return response

	elif param_export_file['format'] == 'bill':

		# print 'DENTROOOOOOOOOOOOO'

		param_export_file['filename'] = param_export_file['filename']+'_' + str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))+'.xlsx'
		filename = param_export_file['filename']
		logger = param_export_file['logger']

		output = StringIO.StringIO()

		book = Workbook(output)
		sheet = book.add_worksheet('test') 

		book.set_properties({
			'title':    param_export_file['filename'],
			'subject':  'Bill',
			'author':   param_export_file['autor'],
			'manager':  'StoneWorkSolutions',
			'company':  'StoneWorkSolutions',
			'category': 'Bill and risk',
			'keywords': 'Bill',
			'comments': 'Created in portal international'})

		response = HttpResponse(writeBillInStream(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
		response['Content-Disposition'] = "attachment; filename="+ filename
		return response




def get_csv_query(request, queryset, filename, col_order = None):

	kind_to_xf_map = {
		'date': ezxf(num_format_str='yyyy-mm-dd HH:MM'),
		'int': ezxf(num_format_str='#,##0'),
		'money': ezxf('font: italic on; pattern: pattern solid, fore-colour grey25',
			num_format_str='€#,##0.00'),
		'price': ezxf(num_format_str='#0.000'),
		'text': ezxf(),
		'boolean': ezxf(),
	}

	try:
		# csv_file = open(filename,"w")
		csv_file = HttpResponse()

		csv_writer = csv.writer(csv_file, dialect='excel', quoting=csv.QUOTE_MINIMAL, delimiter='|')

		if type(queryset)!=list:
			q = queryset[0]

			
			fields = []
			for f in queryset._fields:
				fields.append(q[f])
			for f in queryset.aggregate_names:
				fields.append(q[f])

			fields_name = []
			for f in queryset._fields:
				fields_name.append(unicode(getFieldVerboseName(queryset,f)))
			for f in queryset.aggregate_names:
				fields_name.append(f)


			data_xfs_types_used = [k for k in get_field_types(fields)]
			data_xfs = [kind_to_xf_map[k] for k in get_field_types(fields)]

			data = []


			for i, q in enumerate(queryset):
				aux = []
				i=0
				for f in queryset._fields:

					if data_xfs_types_used[i] == 'date':
						try:
							date_normalize=request['django_timezone'].normalize(q[f]).strftime('%Y-%m-%d %H:%M:%S')
							aux.append(str(date_normalize))
						except:
							# print traceback.format_exc()
							aux.append(q[f])

					else:
						aux.append(q[f])
					i+=1
				for f in queryset.aggregate_names:
					aux.append(q[f])
				data.append(aux)

		else:
			fields_name = []
			if col_order:
				fields_name=col_order
			else:
				for k,v in queryset[0].items():
					fields_name.append(unicode(k))

			data = []

			for q in queryset:
				v_data=[]
				for k in fields_name:
					# print 'excel-->',q[k],'--',k
					v_data.append(q[k])
				data.append(v_data)

		csv_writer.writerow(fields_name)

		for row in data:
			csv_writer.writerow(row)							
								
		# csv_file.close()
		return csv_file

	except Exception, e:
		swslog('error','get_csv_query',e)
		return False


def get_field_types(fields):
	# fields = self.fields
	field_types = []
	prev_len = 0
	found = False
	for lf in fields:
		found = False
		if type(lf) == unicode:
			field_types.append('text')
			found = True
		elif type(lf) == long:
			field_types.append('int')
			found = True
		elif type(lf) == int:
			field_types.append('int')
			found = True
		elif type(lf) == bool:
			field_types.append('boolean')
			found = True
		elif type(lf) == float: 
			field_types.append('price')
			found = True
		elif type(lf) == Decimal: 
			field_types.append('price')
			found = True                    
		elif type(lf) == date:
			field_types.append('date')
			found = True                
		elif type(lf) == datetime: 
			field_types.append('date')
			found = True              
		if found is not True: 
			field_types.append('text')
	return field_types












# def ExportFile_STYLE(param_export_file): # STYLE

# 	if param_export_file['format'] == 'excel':

# 		param_export_file['filename'] = param_export_file['filename']+'_' + str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))+'.xlsx'

# 		queryset = param_export_file['queryset']
# 		headers = param_export_file['headers']
# 		col_name = param_export_file['col_name']
# 		filename = param_export_file['filename']
# 		logger = param_export_file['logger']


# 		# styles = dict(
# 		#     bold = 'font: bold 1',
# 		#     italic = 'font: italic 1',
# 		#     # Wrap text in the cell
# 		#     wrap_bold = 'font: bold 1; align: wrap 1;',
# 		#     # White text on a blue background
# 		#     reversed = 'pattern: pattern solid, fore_color blue; font: color white;',
# 		#     # Light orange checkered background
# 		#     light_orange_bg = 'pattern: pattern fine_dots, fore_color white, back_color orange; font: bold 1; align: wrap 1;',
# 		#     # Heavy borders
# 		#     bordered = 'border: top thick, right thick, bottom thick, left thick;',
# 		#     # 16 pt red text
# 		#     big_red = 'font: height 320, color orange;',
# 		# )



# 		book = xlwt.Workbook()
# 		sheet = book.add_sheet('file_xlsx')
# 		style = xlwt.easyxf('pattern: pattern fine_dots, fore_color white, back_color orange; font: bold 1; align: wrap 1;')
# 		style_content = xlwt.easyxf('font: bold 1')


# 		# wb = Workbook()
# 		# ws = wb.get_active_sheet()
# 		# ws.title = filename


# 		row = 0
# 		col = 0
# 		for h in headers:
# 			sheet.write(row, col, h, style)
# 			# cell = ws.cell(row = row, column = col)
# 			# cell.value = h
# 			# cell.style = style
# 			# printº 'zzzzzzzzzzzzzzzzz',cell
# 			col += 1
	
# 		row = 1
# 		col = 0
# 		for trr in queryset:
# 			for c in col_name:
# 				# cell = ws.cell(row = row, column = col)
# 				sheet.write(row, col, trr[c])
# 				col += 1
# 				# cell.value = trr[c]

				

# 			row += 1
# 			col = 0

# 		# response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') 		
# 		# response['Content-Disposition'] = 'attachment; filename='+filename
# 		# wb.save(response)

# 		# response = HttpResponse(mimetype="application/ms-excel")
# 		# response['Content-Disposition'] = 'attachment; filename= test.xls'
# 		book.save(response)

# 		return response

# 	elif param_export_file['format'] == 'csv':
# 		param_export_file['filename'] = param_export_file['filename']+'_'+str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))

# 		json_data = get_csv_query(param_export_file['request_data'], param_export_file['queryset'], param_export_file['filename'])
# 		response = HttpResponse(json_data, mimetype='text/csv') 
# 		param_export_file['filename'] += '.csv'
# 		response['Content-Disposition'] = 'attachment; filename=' + param_export_file['filename']
# 		return response


# def ExportFileSaveinstream(param_export_file):
# 	if param_export_file['format'] == 'excel':
# 		param_export_file['filename'] = param_export_file['filename']+'_' + str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))+'.xlsx'

# 		queryset = param_export_file['queryset']
# 		headers = param_export_file['headers']
# 		col_name = param_export_file['col_name']
# 		filename = param_export_file['filename']
# 		logger = param_export_file['logger']

# 		response = HttpResponse(mimetype='application/ms-excel') 		
# 		response['Content-Disposition'] = 'attachment; filename='+filename


# 		wb = Workbook()
# 		ws = wb.get_active_sheet()
# 		ws.title = filename

# 		row = 0
# 		col = 0
# 		for h in headers:
# 			cell = ws.cell(row = row, column = col)
# 			cell.value = h
# 			col += 1
	
# 		row = 1
# 		col = 0

# 		wb.save(response)
# 		for trr in queryset:
# 			for c in col_name:
# 				cell = ws.cell(row = row, column = col)
# 				col += 1
# 				cell.value = trr[c]

# 			wb.save(response)
# 			row += 1
# 			col = 0

# 		wb.save(response)
# 		return response